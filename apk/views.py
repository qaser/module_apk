import datetime as dt

from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .filters import FaultFilter
from .forms import FaultForm, FaultFormFirstLevel, FixForm
from .models import Act, Control, Fault, Role
from .utils import check_person, split_on_page


@login_required
def index(request):
    return redirect('apk:index_first_level', slug='1_apk')


@login_required
def index_apk(request, slug):
    if slug == '1_apk':
        return index_first_level(request, slug)
    control = get_object_or_404(Control, slug=slug)
    acts = Act.objects.all().filter(control_level=control)
    # извлекаю уникальные значения годов из выбранных актов и сортирую их
    years_uniq = acts.values('act_year').distinct()
    years = [year.get('act_year') for year in years_uniq]
    context = {}
    # формирую словарь в виде {год: акты этого года}
    for year in years:
        context[year] = acts.filter(act_year=year).order_by('act_number')
    count = len(context)
    return render(
        request,
        'apk/index-acts.html',
        {'years': context, 'control': control, 'count': count},
    )


@login_required
def index_first_level(request, slug):
    control = get_object_or_404(Control, slug=slug)
    if (
        request.user.profile.role != Role.ADMIN and
        request.user.profile.role != Role.MANAGER
    ):
        department = request.user.profile.department
        faults = Fault.objects.all().order_by('-fault_date').filter(
            act__control_level=control,
            location__department=department,
        )
    else:
        faults = Fault.objects.all().order_by('-fault_date').filter(
            act__control_level=control,
        )
    fault_filter = FaultFilter(request.GET, queryset=faults, user=request.user)
    selection = split_on_page(request, fault_filter.qs)
    return render(
        request,
        'apk/index_first_level.html',
        {'control': control, 'fault_filter': fault_filter, **selection}
    )


@login_required
def first_level_single_fault(request, slug, fault_number):
    control = get_object_or_404(Control, slug=slug)
    act = get_object_or_404(
        Act,
        act_year=2021,
        act_number=1,
        control_level=control
    )
    fault = get_object_or_404(
        Fault,
        act=act,
        fault_number=fault_number,
    )
    return render(
        request,
        'apk/single-fault-1-apk.html',
        {'control': control, 'fault': fault}
    )


@login_required
def first_level_fault_new(request, slug):
    control = get_object_or_404(Control, slug=slug)
    act = get_object_or_404(
        Act,
        act_year=2021,
        act_number=1,
        control_level=control
    )
    faults_query = act.faults.all()
    try:
        fault_num = faults_query.latest('fault_number').fault_number + 1
    except ObjectDoesNotExist:
        fault_num = 1
    form = FaultFormFirstLevel(
        request.POST or None,
        files=request.FILES or None,
        user=request.user
    )
    if form.is_valid():
        form.instance.group = 'Охрана труда'
        form.instance.act = act
        form.instance.fault_number = fault_num
        form.instance.inspector = request.user.profile
        form.instance.document = 'Критерии проверки'
        form.save()
        return redirect('apk:index_first_level', slug)
    return render(request, 'apk/form-fault-first-level.html', {'form': form})


@login_required
def single_act(request, slug, act_year, act_number):
    control = get_object_or_404(Control, slug=slug)
    act = get_object_or_404(
        Act,
        act_number=act_number,
        act_year=act_year,
        control_level=control
    )
    faults = Fault.objects.all().filter(act=act)
    return render(
        request,
        'apk/single-act.html',
        {'act': act, 'faults': faults, 'control': control}
    )


@login_required
def single_fault_act(request, slug, act_year, act_number, fault_number):
    control = get_object_or_404(Control, slug=slug)
    act = get_object_or_404(
        Act,
        act_number=act_number,
        act_year=act_year,
        control_level=control
    )
    fault = get_object_or_404(Fault, act=act, fault_number=fault_number)
    return render(
        request,
        'apk/single-fault-act.html',
        {
            'fault': fault,
            'control': control
        }
    )


@login_required
def single_plan(request, slug, act_year, act_number):
    # извлечение "хвоста" запроса из пути
    place = request.GET.getlist('filter')
    control = get_object_or_404(Control, slug=slug)
    act = get_object_or_404(
        Act,
        act_number=act_number,
        act_year=act_year,
        control_level=control
    )
    faults = Fault.objects.all().filter(act=act)
    locations = faults.order_by().values(
        'location__department__title'
    ).distinct()
    places = [i.get('location__department__title') for i in locations]
    # если в "хвосте" запроса есть объект, то фильтрую несоответствия по нему
    if place:
        faults = faults.filter(location__department__title__in=place)
    fixed_faults = faults.filter(fix__fixed=True, fix__corrected=True).count()
    non_fix_faults = 0
    non_correct_faults = 0
    # определение количества просроченных мероприятий
    for fault in faults:
        if (
            fault.fix.deltatime_fix is not None
            and fault.fix.deltatime_fix[1] == 2
        ):
            non_fix_faults = non_fix_faults + 1
        if (
            fault.fix.deltatime_correct is not None
            and fault.fix.deltatime_correct[1] == 2
        ):
            non_correct_faults = non_correct_faults + 1
    return render(
        request,
        'apk/single-plan.html',
        {
            'act': act,
            'faults': faults,
            'filter': place,
            'places': places,
            'control': control,
            'fixed_faults': fixed_faults,
            'non_fix_faults': non_fix_faults,
            'non_correct_faults': non_correct_faults,
        }
    )


@login_required
def single_fault_plan(request, slug, act_year, act_number, fault_number):
    control = get_object_or_404(Control, slug=slug)
    act = get_object_or_404(
        Act,
        act_number=act_number,
        act_year=act_year,
        control_level=control
    )
    fault = get_object_or_404(Fault, act=act, fault_number=fault_number)
    return render(
        request,
        'apk/single-fault-plan.html',
        {
            'fault': fault,
            'control': control
        }
    )


# новый акт
@login_required
def act_new(request, slug):
    # новый акт может добавить только Админ и член ПДК
    if (
        request.user.profile.role != Role.ADMIN and
        request.user.profile.role != Role.MANAGER
    ):
        return redirect('apk:index_control', slug)
    control = get_object_or_404(Control, slug=slug)
    present_year = dt.datetime.today().year
    acts = Act.objects.filter(
        act_year=present_year,
        control_level=control,
    )
    # определяю последний доступный номер акта
    try:
        act_num = acts.latest('act_number').act_number + 1
    except ObjectDoesNotExist:
        act_num = 1
    Act.objects.create(
        control_level=control,
        act_year=present_year,
        act_number=act_num
    )
    return redirect('apk:single_act', slug, present_year, act_num)


# новое несоответствие
@login_required
def fault_new(request, slug, act_year, act_number):
    # новое несоответствие может добавить только Админ и член ПДК
    if (
        request.user.profile.role != Role.ADMIN and
        request.user.profile.role != Role.MANAGER
    ):
        return redirect('apk:single_act', slug, act_year, act_number)
    control = get_object_or_404(Control, slug=slug)
    act = get_object_or_404(
        Act,
        act_year=act_year,
        act_number=act_number,
        control_level=control
    )
    faults_query = act.faults.all()
    try:
        fault_num = faults_query.latest('fault_number').fault_number + 1
    except ObjectDoesNotExist:
        fault_num = 1
    form = FaultForm(request.POST or None, files=request.FILES or None)
    if form.is_valid():
        form.instance.act = act
        form.instance.fault_number = fault_num
        form.instance.inspector = request.user.profile
        form.save()
        return redirect('apk:single_act', slug, act_year, act_number)
    return render(request, 'apk/form-fault.html', {'form': form})


# редактировать несоответствие
@login_required
def fault_edit(request, slug, act_year, act_number, fault_number):
    control = get_object_or_404(Control, slug=slug)
    act = get_object_or_404(
        Act,
        act_year=act_year,
        act_number=act_number,
        control_level=control
    )
    fault = get_object_or_404(Fault, fault_number=fault_number, act=act)
    # редактировать несоответствие может только автор или админ
    if (
        request.user.profile != fault.inspector and
        request.user.profile.role != Role.ADMIN
    ):
        return redirect(
            'apk:single_fault_act',
            slug,
            act_year,
            act_number,
            fault_number
        )
    form = FaultForm(
        request.POST or None,
        files=request.FILES or None,
        instance=fault
    )
    if form.is_valid():
        form.save()
        return redirect(
            'apk:single_fault_act',
            slug,
            act_year,
            act_number,
            fault_number
        )
    return render(
        request,
        'apk/form-fault.html',
        {'form': form, 'fault': fault}
    )


# новые мероприятия
@login_required
def fix_new(request, slug, act_year, act_number, fault_number):
    if request.user.profile.role == Role.EMPLOYEE:
        return redirect(
            'apk:single_fault_plan',
            slug,
            act_year,
            act_number,
            fault_number
        )
    control = get_object_or_404(Control, slug=slug)
    act = get_object_or_404(
        Act,
        act_year=act_year,
        act_number=act_number,
        control_level=control
    )
    fault = get_object_or_404(Fault, fault_number=fault_number, act=act)
    form = FixForm(
        request.POST or None,
        files=request.FILES or None,
        instance=fault.fix
    )
    if form.is_valid():
        form.save()
        return redirect(
            'apk:single_fault_plan',
            slug,
            act_year,
            act_number,
            fault_number
        )
    return render(
        request,
        'apk/form-fix.html',
        {'form': form, 'fault': fault}
    )


@login_required
def export_act_excel(request, slug, act_year, act_number):
    control = get_object_or_404(Control, slug=slug)
    act = get_object_or_404(
        Act,
        act_number=act_number,
        act_year=act_year,
        control_level=control
    )
    faults = Fault.objects.all().filter(act=act)
    response = HttpResponse(
        content_type=('application/vnd.openxmlformats-officedocument.'
                      'spreadsheetml.sheet'),
    )
    filename = 'act.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    workbook = Workbook()
    worksheet = workbook.active
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    worksheet.title = f'Акт {act_number} комплексной проверки'
    columns = [
        ('№ п/п', 5),
        ('Тема технической учёбы', 60),
        ('№ вопроса', 10),
        ('Содержание вопроса', 80),
        ('Варианты ответов', 80),
    ]
    row_num = 1
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width
        worksheet.row_dimensions[row_num].height = 110
    for fault in faults:
        row_num += 1
        row = [
            fault.fault_number,
            '{}; {}. {}'.format(
                fault.location.department,
                fault.location.object,
                fault.description
            ),
            fault.document,
            fault.danger_readable,
            fault.intruder.lastname_and_initials,
            fault.unseeing.lastname_and_initials,
            fault.section_esupb,
            fault.inspector.lastname_and_initials,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            worksheet.row_dimensions[row_num].height = 50
            cell.value = cell_value
            cell.alignment = wrapped_alignment
    workbook.save(response)
    return response


@login_required
def export_plan_excel(request, slug, act_year, act_number):
    control = get_object_or_404(Control, slug=slug)
    act = get_object_or_404(
        Act,
        act_number=act_number,
        act_year=act_year,
        control_level=control
    )
    faults = Fault.objects.all().filter(act=act)
    response = HttpResponse(
        content_type=('application/vnd.openxmlformats-officedocument'
                      '.spreadsheetml.sheet'),
    )
    filename = 'plan.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    workbook = Workbook()
    worksheet = workbook.active
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    worksheet.title = ('План корректирующих действий '
                       f'согласно Акту №{act_number}')
    columns = [
        ('№ п/п', 5),
        ('Описание несоответствия, пункт НТД', 60),
        ('Наименование мероприятия', 40),
        ('Ответственный исполнитель (подпись, дата)', 15),
        ('Срок выполнения', 20),
        ('Причины появления несоответствия', 20),
        ('Корректирующее действие (мероприятие)', 20),
        ('Требуемые условия и ресурсы', 20),
        ('Ответственный исполнитель (подпись, дата)', 20),
        ('Срок выполнения', 20)
    ]
    row_num = 1
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width
        worksheet.row_dimensions[row_num].height = 70
    for fault in faults:
        row_num += 1
        row = [
            fault.fault_number,
            '{}; {}. {}. {}'.format(
                fault.location.department,
                fault.location.object,
                fault.description,
                fault.document
            ),
            fault.fix.fix_action,
            check_person(fault.fix.fixer),
            fault.fix.fix_deadline,
            fault.fix.reason,
            fault.fix.correct_action,
            fault.fix.resources,
            check_person(fault.fix.corrector),
            fault.fix.correct_deadline
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            worksheet.row_dimensions[row_num].height = 50
            cell.value = cell_value
            cell.alignment = wrapped_alignment
    workbook.save(response)
    return response
